'''
- Entrar na planilha e extrair o cpf do cliente.
- Entro no site https://consultcpf-devaprender.netlify.app/ e uso o cpf da planilha para pesquisar o status do pagamento.
- Verificar se está "em dia" ou "atrasado".
- Se estiver "em dia", pegar a data do pagamento e o método de pagamento.
- Caso contrário (se estiver atrasado), colocar o status como pendente.
- Inserir essas novas informações (nome, valor, cpf, vencimento, status e caso esteja em dia, data pagamento, método
pagamento do pagamento(boleto ou cartão) em uma nova planilha.
- Repetir até chegar no último cliente.

'''

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# Função para obter status de pagamento
def obter_status_pagamento(driver, cpf):
    driver.get('https://consultcpf-devaprender.netlify.app/')
    sleep(3)

    campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    sleep(1)

    botao_pesquisar = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    botao_pesquisar.click()
    sleep(4)

    try:
        status_element = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
        status = status_element.text.strip()
    except Exception as e:
        print(f"Erro ao obter status para CPF {cpf}: {e}")
        status = ""

    return status

# Carregar a planilha de clientes
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

# Iniciar o driver do Selenium
driver = webdriver.Chrome()

# Carregar a planilha de fechamento
planilha_fechamento = openpyxl.Workbook()
aba_fechamento = planilha_fechamento.active
aba_fechamento.title = 'Sheet'  # Definir o título da aba como 'Sheet'

# Adicionar cabeçalho na planilha de fechamento
aba_fechamento.append(['Nome', 'Valor', 'CPF', 'Vencimento', 'Status', 'Data Pagamento', 'Método Pagamento'])

# Processar cada linha da planilha de clientes
for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha
    
    # Obter o status de pagamento usando a função definida
    status = obter_status_pagamento(driver, cpf)

    # Processar o status e obter dados adicionais, se aplicável
    if status == "em dia":
        try:
            data_pagamento_element = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
            data_pagamento = data_pagamento_element.text.split()[3] if data_pagamento_element else ""
        except Exception as e:
            print(f"Erro ao obter data de pagamento para CPF {cpf}: {e}")
            data_pagamento = ""

        try:
            metodo_pagamento_element = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")
            metodo_pagamento = metodo_pagamento_element.text.split()[3] if metodo_pagamento_element else ""
        except Exception as e:
            print(f"Erro ao obter método de pagamento para CPF {cpf}: {e}")
            metodo_pagamento = ""

        # Adicionar dados à planilha de fechamento
        aba_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento, metodo_pagamento])
    else:
        # Caso contrário, marcar como pendente
        aba_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])

# Salvar a planilha de fechamento
try:
    planilha_fechamento.save('planilha_report.xlsx')
    print("Dados salvos com sucesso na planilha de fechamento.")
except Exception as e:
    print(f"Erro ao salvar a planilha de fechamento: {e}")

# Fechar o driver do Selenium
driver.quit()







    