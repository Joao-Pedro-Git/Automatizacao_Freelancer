'''
- Entrar na planilha e extrair o cpf do cliente.
- Entro no site https://consultcpf-devaprender.netlify.app/ e uso o cpf da planilha para pesquisar o status do pagamento.
- Verificar se está "em dia" ou "atrasado".
- Se estiver "em dia", pegar a data do pagamento e o método de pagamento.
- Caso contrário (se estiver atrasado), colocar o status como pendente.
- Inserir essas novas informações (nome, valor, cpf, vencimento, status e caso esteja em dia, data pagamento, método
pagamento do pagamento(boleto ou cartão) em uma nova planilha.
- Repetir até chegar no último cliente.

'''

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# Entrar na planilha e extrair o cpf do cliente
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')

# Entrar no site https://consultcpf-devaprender.netlify.app/ e usar o cpf da planilha para pesquisar o status do pagamento.
driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')
sleep(5)

pagina_clientes = planilha_clientes['Sheet1']

# Estrutura de Loop com For

for linha in pagina_clientes.iter_rows(min_row=2, values_only=True): # Qual linha ele deve iniciar(ignore a primeira por isso rows=2)
    nome, valor, cpf, vencimento = linha
    
    campo_pesquisa = driver.find_element(By.XPATH,"//input[@id='cpfInput']")

    sleep(1)

    campo_pesquisa.clear() # Apagar para color outro cpf
    campo_pesquisa.send_keys(cpf) 

    sleep(1)

    botao_pesquisar = driver.find_element(By.XPATH,"//button[@class='btn btn-custom btn-lg btn-block mt-3']")

    sleep(2)                          

    botao_pesquisar.click()
    sleep(3)


    # Verificar se está "em dia" ou "atrasado".
    status = driver.find_element(By.XPATH,"//span[@id='statusLabel']")
    status.text
    

    if status.text == "em dia":

        # Se estiver "em dia", pegar a data do pagamento e o método de pagamento.

        data_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentMethod']")

        data_pagamento_limpo = data_pagamento.text.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]

        planilha_fechamento = openpyxl.load_workbook('planilha_report.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome,valor,cpf,vencimento,'em dia',data_pagamento_limpo,metodo_pagamento_limpo])

        planilha_fechamento.save('planilha_report.xlsx')
        print("Dados salvos com sucesso na planilha de fechamento.")
        
    else:

        # Caso contrário (se estiver atrasado), colocar o status como pendente.
        planilha_fechamento = openpyxl.load_workbook('planilha_report.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome,valor,cpf,vencimento,'pendente'])

        planilha_fechamento.save('planilha_report.xlsx')
        print("Dados salvos com sucesso na planilha de fechamento.")

# Fechar o driver do Selenium
driver.quit()
